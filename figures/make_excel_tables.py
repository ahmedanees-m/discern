"""Export the large supplemental tables as formatted Excel workbooks.

Cell Press cannot typeset a CSV: a comma-separated file is a data file, not a display item. Tables
S3 (every discrimination likelihood ratio with its source), S5 (the curated case set) and S7 (the
data source manifest) are legitimately large data tables, which the journal permits as separate
Excel files. The CSVs stay in the Zenodo deposit as the machine-readable archive; these are the
journal artifacts.

Run:  python -m figures.make_excel_tables [--csv DIR] [--out DIR]
"""
from __future__ import annotations

import argparse
import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(__file__)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(size=10)
NOTE_FONT = Font(size=9, italic=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# csv stem -> (workbook name, sheet title, caption placed above the table)
EXPORTS = {
    "tableS3_cluster_likelihood_ratios": (
        "Table_S3", "S3 likelihood ratios",
        "Table S3. Discrimination likelihood ratios for the ten confusable-disease clusters. "
        "Each value is the probability of the finding given the disease, with the sample size and "
        "the PubMed identifier of the primary source from which it was taken. Continuous "
        "integration fails if any entry lacks a source."),
    "tableS5_curated_cases": (
        "Table_S5", "S5 curated cases",
        "Table S5. The curated published-case benchmark (n = 42). Each case is represented by its "
        "PubMed identifier, causal gene, cluster, expected diagnosis and extracted Human Phenotype "
        "Ontology terms, together with the rank each method assigned. No article text is "
        "reproduced."),
    "tableS7_data_source_manifest": (
        "Table_S7", "S7 data sources",
        "Table S7. Third-party data sources, with the files used, the version or snapshot date and "
        "the role each plays. None is redistributed; checksums for the exact files used are in the "
        "deposit manifest."),
    "tableS9_software_versions": (
        "Table_S9", "S9 software versions",
        "Table S9. Software, tool and database versions used to produce every reported result."),
}


def _rows(path):
    with open(path, encoding="utf-8") as fh:
        return list(csv.reader(fh))


def export(csv_path, out_dir, book, sheet, caption):
    rows = _rows(csv_path)
    header, body = rows[0], [r for r in rows[1:] if r and any(c.strip() for c in r)]
    notes = [r[0] for r in body if r and r[0].startswith("NOTE:")]
    body = [r for r in body if not (r and r[0].startswith("NOTE:"))]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet[:31]

    ws.cell(row=1, column=1, value=caption).font = Font(size=10, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(header), 2))
    ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 46

    for j, h in enumerate(header, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.fill, c.font, c.border = HEADER_FILL, HEADER_FONT, BORDER
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[3].height = 30

    for i, row in enumerate(body, start=4):
        for j, val in enumerate(row, start=1):
            try:
                v = float(val) if val not in ("", None) and val.strip("-").replace(".", "", 1).isdigit() else val
            except (ValueError, AttributeError):
                v = val
            c = ws.cell(row=i, column=j, value=v)
            c.font, c.border = BODY_FONT, BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")

    r = len(body) + 5
    for n in notes:
        ws.cell(row=r, column=1, value=n).font = NOTE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(len(header), 2))
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 34
        r += 1

    for j, h in enumerate(header, start=1):
        longest = max([len(str(h))] + [len(str(row[j - 1])) for row in body if len(row) >= j])
        ws.column_dimensions[get_column_letter(j)].width = min(max(12, longest + 2), 58)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(header))}{len(body) + 3}"

    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"{book}.xlsx")
    wb.save(path)
    return path, len(body)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", default=os.path.join(HERE, "tables"))
    ap.add_argument("--out", default=os.path.join(HERE, "excel"))
    args = ap.parse_args()
    for stem, (book, sheet, caption) in EXPORTS.items():
        src = os.path.join(args.csv, f"{stem}.csv")
        if not os.path.exists(src):
            print(f"  ! missing {stem}.csv")
            continue
        path, n = export(src, args.out, book, sheet, caption)
        print(f"  {book:<10} {n:>4} rows -> {os.path.basename(path)}")
    print(f"\nwrote Excel tables to {args.out}")


if __name__ == "__main__":
    main()
